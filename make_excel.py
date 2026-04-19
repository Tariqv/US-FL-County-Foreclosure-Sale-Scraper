import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scraper import scrape_county
from upcoming import find_next_upcoming
from db.base_url import COUNTY_URLS
from utils import get_auction_date, extract_from_address, SESSION, init_usaddress
import os
from curl_cffi.requests.exceptions import RequestException

EXPECTED_MAP = {
    "county": "County",
    "auction sold": "Auction Sold",
    "case": "Case #",
    "parcel id": "Parcel ID",
    "property address": "Property Address",
    "City": "City",
    "State": "State",
    "Zip": "Zip",
    "final judgment amount": "Final Judgment Amount",
    "amount": "Amount",
    "sold_to": "Sold To",
    "auction type": "Auction Type",
    "auction time": "Auction Time",
    "upcoming date": "Upcoming Date",
}

def check_404_status():
    session = SESSION()
    url = 'https://www.alachua.realforeclose.com/'
    
    r = session.get(url)
    if r.status_code in (404, 403):
        return True
    return False

def norm(x: str):
    return x.lower().strip().replace(":", "").replace("#", "").replace("_", " ")

def build_rows(county, auctions):
    rows = []
    for a in auctions:
        only_address, City, State, Zip = extract_from_address(a.get("property_address"))
        rows.append({
            "County": county.upper(),
            "Auction Sold": a.get("auction_time"),
            "Case #": a.get("case_number"),
            "Parcel ID": a.get("parcel_id"),
            "Property Address": only_address,
            "City": City,
            "State": State,
            "Zip": Zip,
            "Final Judgment Amount": a.get("final_judgment"),
            "Amount": a.get("amount"),
            "Sold To": a.get("sold_to").upper(),
            "Auction Type": a.get("auction_type") or "FORECLOSURE",
        })

    return rows


DEFAUTL_PATH = 'FL Forclosure Final Report'
DEFAUTL_AUCTION_DATE = get_auction_date()
def main(output_path=DEFAUTL_PATH, auction_date=DEFAUTL_AUCTION_DATE):
    init_usaddress()
    all_rows = []
    sheet2_rows = []

    if check_404_status():
        print('Please Reconnect or Change loaction from Vpn.')
        print('Thanks For Using...')
        return

    print("👋 Welcome to FL Foreclosure County Scraper...\n")
    print(f"=== Start Scraping {auction_date} ===\n")

    os.makedirs(output_path, exist_ok=True)
    output_file =  os.path.join(output_path, f"Final_{auction_date.replace('/', '-')}.xlsx")
    if os.path.exists(output_file):
        try:
            os.rename(output_file, output_file)
            print('FILE ALREADY EXISTS RE SCRAPING AND UPDATING DATA.')
        except PermissionError:
            print("⚠️ Please close the Excel file first and rerun application.")
            return

    for county, base_url in COUNTY_URLS.items():
        auctions = scrape_county(county, base_url, auction_date)
        if auctions:
            rows = build_rows(county, auctions)
            all_rows.extend(rows)
        else:
            print(f'⚠️ Skipping {county} No Data Found.')
        upcoming = find_next_upcoming(base_url, auction_date)
        if not upcoming:
            print(f'⚠️ No upcoming auction found for {county} in 12 months.')
        sheet2_rows.append({
            "County": county.upper(),
            "Upcoming Date": upcoming.get('next_date') or "No upcoming auction found in 12 months for this county",
            "Auction Type": "FORECLOSURE"
        })
    df2 = pd.DataFrame(sheet2_rows)
    if all_rows:
        df = pd.DataFrame(all_rows)

        normalized = {
            col: EXPECTED_MAP[norm(col)]
            for col in df.columns
            if norm(col) in EXPECTED_MAP
        }

        df = df.rename(columns=normalized)

        final_cols = list(EXPECTED_MAP.values())
        df = df[[c for c in final_cols if c in df.columns]]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)
            df2.to_excel(writer, sheet_name="Sheet2", index=False)

    else:
        print("⚠️ No 3rd party bidder auction found. Writing upcoming auctions on the First Sheet.")
        df2.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save(output_file)

    print("\n✅ DONE Saved in ", output_file)
    print('Thanks For Using...')

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('Interrupted by User')
        sys.exit(32)
    except RequestException as e:
        print(f'Request failed {e}')
        sys.exit(1)