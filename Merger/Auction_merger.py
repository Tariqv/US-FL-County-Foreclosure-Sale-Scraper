import os
from datetime import datetime, timedelta


def normalize_column(col_name):
    return (
        col_name.strip()
        .lower()
        .replace(":", "")
        .replace("#", "")
        .replace("  ", " ")
        .strip()
    )


def extract_from_address(address):
    import pandas as pd
    import re

    address = str(address).strip().upper()

    # 1️⃣ Extract ZIP from the end
    zip_match = re.search(r"(\d{5})$", address)
    if not zip_match:
        return pd.Series(["", "", "", ""])

    zipcode = zip_match.group(1)
    state = "FL"  # Default to FL

    # 2️⃣ Remove ZIP from end
    address_before_zip = re.sub(r"[\s,-]*\d{5}$", "", address).strip()

    # 3️⃣ Remove trailing FL or FL- if present
    address_before_zip = re.sub(r"(,?\s*FL-?\s*)$", "", address_before_zip).strip()

    # 4️⃣ Split by comma
    parts = [p.strip() for p in address_before_zip.split(",") if p.strip()]

    if len(parts) >= 2:
        city = parts[-1]
        address_only = ", ".join(parts[:-1])
    else:
        city = ""
        address_only = parts[0] if parts else ""

    # Ensure uppercase & trimmed
    city = city.upper().strip()
    address_only = address_only.upper().strip()
    return pd.Series([state, city, zipcode, address_only])


def main():
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import pytz

    real_date = pytz.timezone("Asia/Kolkata")
    ist = datetime.now(pytz.timezone("Asia/Kolkata")).replace(day=1) - timedelta(days=1)
    now_ist = datetime.now(real_date)
    target_date_str = ist.strftime("%m-%d-%Y")
    yesterday = datetime.now() - timedelta(days=1)
    if now_ist.day == 1:
        auction_date = target_date_str
    else:
        auction_date = yesterday.strftime("%m/%d/%Y")
    folder_name = auction_date.replace("/", "-")

    if not os.path.exists(folder_name):
        print(f"⚠️ Folder '{folder_name}' not found!")
        return

    # Target column names in final output
    desired_columns = [
        "County",
        "Auction Sold",
        "Case #",
        "Parcel ID",
        "Property Address",
        "City",
        "State",
        "Zip",
        "Final Judgment Amount",
        "Amount",
        "Sold To",
        "Auction Type",
    ]

    # Normalized key map for fuzzy matching
    expected_map = {
        "county": "County",
        "auction sold": "Auction Sold",
        "case": "Case #",
        "case number": "Case #",
        "parcel id": "Parcel ID",
        "property address": "Property Address",
        "final judgment amount": "Final Judgment Amount",
        "amount": "Amount",
        "sold to": "Sold To",
        "auction type": "Auction Type",
    }

    all_data = []

    print("\n\n\n=== Merging Data ===")
    print(f"\n🔄 Reading Excel files from: '{folder_name}'")

    for file in os.listdir(folder_name):
        if file.endswith(".xlsx"):
            file_path = os.path.join(folder_name, file)
            try:
                df = pd.read_excel(file_path)

                # Build normalized column mapping
                normalized_map = {normalize_column(col): col for col in df.columns}

                # Remap columns to standard names
                col_mapping = {}
                for norm_col, orig_col in normalized_map.items():
                    if norm_col in expected_map:
                        col_mapping[orig_col] = expected_map[norm_col]

                df = df.rename(columns=col_mapping)

                # Skip if "Sold To" column doesn't exist
                if "Sold To" not in df.columns:
                    print(f"⚠️ Skipping '{file}' — 'Sold To' column not found")
                    continue

                filtered_df = df[
                    df["Sold To"].astype(str).str.strip().str.lower()
                    == "3rd party bidder"
                ]
                if (
                    filtered_df["Parcel ID"]
                    .astype(str)
                    .str.upper()
                    .str.contains("TIMESHARE")
                    .any()
                ):
                    print(
                        f"⛔ Found 'TIMESHARE' rows in '{file}' — skipping this rows."
                    )
                    filtered_df = filtered_df[
                        ~filtered_df["Parcel ID"]
                        .astype(str)
                        .str.upper()
                        .str.contains("TIMESHARE")
                    ]

                if filtered_df.empty:
                    print(f"⚠️ No '3rd Party Bidder' found in: {file}")
                    continue  # ⛔ Don't add any dummy row

                # Only process if actual 3rd party bidders are found
                new_cols_df = (
                    filtered_df["Property Address"]
                    .apply(extract_from_address)
                    .apply(pd.Series)
                )
                new_cols_df.columns = ["State", "City", "Zip", "Clean Address"]
                new_cols_df = new_cols_df.apply(
                    lambda col: col.map(
                        lambda x: str(x).upper() if isinstance(x, str) else x
                    )
                )
                filtered_df.loc[:, ["State", "City", "Zip"]] = new_cols_df[
                    ["State", "City", "Zip"]
                ]
                filtered_df.loc[:, "Property Address"] = new_cols_df["Clean Address"]

                # Add 'County' from file name
                county_name = file.split("_")[0].upper()
                filtered_df.insert(0, "County", county_name)

                # Ensure all desired columns exist
                for col in desired_columns:
                    if col not in filtered_df.columns:
                        filtered_df[col] = ""

                # Reorder columns
                filtered_df = filtered_df[desired_columns]

                all_data.append(filtered_df)
                print(f"✅ Added {len(filtered_df)} rows from '{file}'")

            except Exception as e:
                print(f"⚠️ Error reading '{file}': {e}")

    if not all_data:
        print("⚠️ No data for 3rd party sale.")

        # 🧹 Still try to delete files even if no merge
        try:
            import shutil

            input = f"availability_of_{folder_name}.xlsx"
            if os.path.exists(input):
                os.remove(input)
                print(f"✅ Deleted file '{input}'")

            if os.path.exists(folder_name) and os.path.isdir(folder_name):
                shutil.rmtree(folder_name)
                print(f"✅ Deleted folder '{folder_name}'")
        except Exception as cleanup_error:
            print(f"⚠️ Cleanup failed: {cleanup_error}")

        print("✅ Thank You.")
        return

    print("🔄 Merging data...")

    final_df = pd.concat(all_data, ignore_index=True)
    # Fill missing or blank Auction Type values with 'Foreclosure'

    final_df["Auction Type"] = final_df["Auction Type"].apply(
        lambda x: "FORECLOSURE" if pd.isna(x) or str(x).strip() == "" else x
    )

    # Create output folder
    output_folder = "FL Forclosure Final Report"
    os.makedirs(output_folder, exist_ok=True)

    # Save final file inside the folder
    output_file = os.path.join(output_folder, f"Final_{folder_name}.xlsx")
    final_df.to_excel(output_file, index=False)
    print(f"✅ Merged file created: '{output_file}'")
    wb = load_workbook(output_file)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)
    try:
        input = f"availability_of_{folder_name}" + ".xlsx"
        df = pd.read_excel(input)

        # 2. Filter the columns you need
        filtered_data = df[["County", "Upcoming Date", "Auction Type"]]

        # 3. Target file (where Sheet1 + Sheet2 will exist)
        output_file = f"FL Forclosure Final Report/Final_{folder_name}.xlsx"

        # 4. Load or create the target workbook
        try:
            # Try loading existing workbook
            book = load_workbook(output_file)
        except FileNotFoundError:
            # If file doesn't exist, create a new workbook with Sheet1
            book = load_workbook()
            book.save(output_file)
            book = load_workbook(output_file)

        # 5. Remove Sheet2 if it already exists
        if "Sheet2" in book.sheetnames:
            book.remove(book["Sheet2"])

        # 6. Create a new Sheet2 and write data
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
            mode="a",  # Append mode (preserves existing sheets)
        ) as writer:
            filtered_data.to_excel(writer, sheet_name="Sheet2", index=False)

        try:
            import shutil

            input = f"availability_of_{folder_name}.xlsx"
            if os.path.exists(input):
                os.remove(input)
                print(f"✅ Deleted file '{input}'")

            if os.path.exists(folder_name) and os.path.isdir(folder_name):
                shutil.rmtree(folder_name)
                print(f"✅ Deleted folder '{folder_name}'")
        except Exception as cleanup_error:
            print(f"⚠️ Cleanup failed: {cleanup_error}")
        print("✅ Thank You.")
        return
    except Exception as e:
        print("❌ Error during Sheet2 creation:", e)
        # 🧹 Still try to delete files even if no merg


if __name__ == "__main__":
    main()
